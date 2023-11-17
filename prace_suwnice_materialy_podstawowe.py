import pandas as pd
import openpyxl
from tkinter import filedialog
import os
import subprocess
import re
import customtkinter


def materialy_podstawowe():
    customtkinter.set_appearance_mode("dark")
    customtkinter.set_default_color_theme("blue")
    root = customtkinter.CTk()

    root.title("Załaduj Prace Suwnice")
    root.geometry("300x155")
    root.resizable(0, 0)

    arkusz = ()
    def search():
        global arkusz
        arkusz = entry.get()
        temp = list(arkusz)
        temp.clear()
        clear_entry()
        realizer()
        entry_label.destroy()
        new_label = customtkinter.CTkLabel(root, text=("Prace suwnice zostały pomyślnie \n zaktualizowane"), text_color=("#2AAA8A", "#2AAA8A"))
        new_label.grid(row=0, column=0, columnspan=1, padx=30, pady=25, sticky="n")

    def clear_entry():
        return entry.delete(0, 'end')

    entry = customtkinter.CTkEntry(master=root, width=130, height=35, border_width=1, justify="center")
    entry.grid(row=0, column=0, columnspan=1, padx=90, pady=62, sticky="w")

    entry_label = customtkinter.CTkLabel(root, text=("Wprowadz nazwę arkusza np. 11-2023, \n a następnie zatwierdź"))
    entry_label.grid(row=0, column=0, columnspan=1, padx=30, pady=25, sticky="n")

    search_button = customtkinter.CTkButton(root, text='Zatwierdz', width=100, height=35, border_width=1, command=search)
    search_button.grid(row=0, column=0, padx=15, pady=20, sticky="s")
    entry.bind("<Return>", lambda event: search())


    def realizer():
        global arkusz
        miesiac = arkusz

        try:
            path = r'C:\Users\Marcin\Desktop\Prace 2023.xlsx'
            df = pd.read_excel(path, sheet_name=miesiac, header=0)
        except FileNotFoundError:
            path = filedialog.askopenfilename(initialdir="/", title="Załaduj prace suwnice",
                                          filetypes=(("pliki excel", "*.xlsx"), ("wszystkie pliki", "*")))
            df = pd.read_excel(path, sheet_name=miesiac, header=0)

        df_baza = pd.read_excel('dfall.xlsx')
        df_baza = df_baza.loc[(df_baza['Data'] >= f'2022-06-01')]
        df_baza = df_baza[['Data', 'Material', 'Cena']]
        df_baza = df_baza.sort_values(by='Data').drop_duplicates('Material', keep='last')
        df_baza['Material'] = df_baza['Material'].str.strip()
        df_baza['Material'] = df_baza['Material'].replace(regex=[r'- '], value='')

        workbook = openpyxl.load_workbook(path)

        worksheet = workbook[miesiac]

        # Dodawanie cen na podstawie nazw materiałów, formatowanie tekstu
        for row in worksheet.iter_rows():
            if row[3].value !=None:
                row[3].value = row[3].value.replace('taśma izolacyjna', 'REPERO taśma 19x20 czarna')
                row[3].value = row[3].value.replace('tasma izolacyjna', 'REPERO taśma 19x20 czarna')
                row[3].value = row[3].value.replace('Taśma 19x20 CZARNA', 'REPERO taśma 19x20 czarna')
                row[3].value = row[3].value.replace(' ,', ',')
                row[3].value = row[3].value.replace(' , ', ', ')
                row[3].value = row[3].value.replace('  ,', ',')
                row[3].value = row[3].value.replace('-', ' - ')
                row[3].value = row[3].value.replace('   -', ' -')
                row[3].value = row[3].value.replace('  -', ' -')
                row[3].value = row[3].value.replace('-   ', '- ')
                row[3].value = row[3].value.replace('-  ', '- ')
                row[3].value = row[3].value.replace('szt.', 'szt')
                row[3].value = row[3].value.replace('rbh.', 'rbh')
                row[3].value = row[3].value.replace('kg.', 'kg')
                row[3].value = row[3].value.replace('op.', 'kg')
                row[3].value = row[3].value.replace(',,', ',')
                row[3].value = row[3].value.replace(' .', '.')
                row[3].value = row[3].value.replace('  ', ' ')
                row[3].value = row[3].value.replace(' ;', ';')
                row[3].value = row[3].value.replace('( ', '(')
                row[3].value = row[3].value.replace(' )', ')')
                row[3].value = row[3].value.replace('F55 - 34 - 9 - 024 - 004', 'F55-34-9-024-004')
                row[3].value = row[3].value.replace('WT - ', 'WT-')
                row[3].value = row[3].value.replace('WTS - ', 'WTS-')
                row[3].value = row[3].value.replace('WD - ', 'WD-')
                row[3].value = row[3].value.replace('WTs - ', 'WTs-')
                row[3].value = row[3].value.replace('BTP - ', 'BTP-')
                row[3].value = row[3].value.replace(' - GG - ', '-GG-')
                row[3].value = row[3].value.replace('gG - ', 'gG-')
                row[3].value = row[3].value.replace('  ', ' ')
                for material, cena in zip(df_baza['Material'], df_baza['Cena']):
                    if material in row[3].value:
                        row[6].value = cena
                        row[6].number_format = '#,##0.00 zł'

        # Przeliczenie materiałów typu śruby, nakładki, podkładki - szt na kg
        pattern = r'[-][ ]([0-9]+.{1,}|[0-9])+ (?:szt|kpl|op|kg|mb|l|m|m3|km)'
        match = ()

        def sruby_szt_kg():
            try:
                match = row[3].value
                ile = re.findall(pattern, match)
                print(ile)
                szt = float(ile[0])
                kg = round((szt / 1000) * stala, 2)
                szt_str = str(ile[0])
                kg_str = str(kg)
                row[3].value = row[3].value.replace(f'{szt_str} szt', f'{kg_str} kg')
                row[3].value = row[3].value.replace('.', ',')
            except IndexError:
                pass
            except ValueError:
                pass

        for row in worksheet.iter_rows():
            if row[3].value is not None:
                if 'śruba' in row[3].value:
                    if 'szt' in row[3].value:
                        if '20*50' in row[3].value:
                            stala = 175.27
                            sruby_szt_kg()
                        elif '20*60' in row[3].value:
                            stala = 195.75
                            sruby_szt_kg()
                        elif '16*40' in row[3].value:
                            stala = 88.72
                            sruby_szt_kg()
                        elif '16*30' in row[3].value:
                            stala = 75.76
                            sruby_szt_kg()
                        elif '16*80' in row[3].value:
                            stala = 140.54
                            sruby_szt_kg()
                        elif '12*70' in row[3].value:
                            stala = 66.81
                            sruby_szt_kg()
                        elif '12*40' in row[3].value:
                            stala = 45.65
                            sruby_szt_kg()
                        elif '12*60' in row[3].value:
                            stala = 59.76
                            sruby_szt_kg()
                        elif '12*100' in row[3].value:
                            stala = 87.98
                            sruby_szt_kg()
                        elif '12*120' in row[3].value:
                            stala = 102.07
                            sruby_szt_kg()
                        elif '8*60' in row[3].value:
                            stala = 24.17
                            sruby_szt_kg()
                        elif '24*110' in row[3].value:
                            stala = 450.33
                            sruby_szt_kg()
                        elif '10*30' in row[3].value:
                            stala = 26.47
                            sruby_szt_kg()
                        elif '10*40' in row[3].value:
                            stala = 31.34
                            sruby_szt_kg()
                elif 'nakrętka' in row[3].value:
                    if 'szt' in row[3].value:
                        if 'M6' in row[3].value:
                            stala = 2.46
                            sruby_szt_kg()
                        elif 'M8' in row[3].value:
                            stala = 5.31
                            sruby_szt_kg()
                        elif 'M10' or 'M 10' in row[3].value:
                            stala = 9.84
                            sruby_szt_kg()
                        elif 'M12' or 'M 12' in row[3].value:
                            stala = 15
                            sruby_szt_kg()
                        elif 'M16' or 'M 16' in row[3].value:
                            stala = 36.2
                            sruby_szt_kg()
                        elif 'M20' or 'M 20' in row[3].value:
                            stala = 69.1
                            sruby_szt_kg()
                        elif 'M24' or 'M 24' in row[3].value:
                            stala = 117
                            sruby_szt_kg()
                elif 'podkładka' in row[3].value:
                    if 'szt' in row[3].value:
                        if 'fi 8' in row[3].value:
                            stala = 2.15
                            sruby_szt_kg()
                        if 'fi 10' or 'M 10' in row[3].value:
                            stala = 4.08
                            sruby_szt_kg()
                        if 'fi 12' or 'M 12' in row[3].value:
                            stala = 6.27
                            sruby_szt_kg()
                        if 'fi 16' or 'M 16' in row[3].value:
                            stala = 11.3
                            sruby_szt_kg()
                        if 'fi 20' in row[3].value:
                            stala = 17.1
                            sruby_szt_kg()
                        if 'fi 24' in row[3].value:
                            stala = 32.3
                            sruby_szt_kg()
                        if 'fi 6' in row[3].value:
                            stala = 1.13
                            sruby_szt_kg()

        workbook.save(path)

        # Baza i ukrycie bazy
        if os.path.exists("baza_materialy_podstawowe.xlsx"):
            os.remove("baza_materialy_podstawowe.xlsx")
        df_baza.to_excel("baza_materialy_podstawowe.xlsx", index=False)
        subprocess.check_call(["attrib", "+H", "baza_materialy_podstawowe.xlsx"])

    root.mainloop()
