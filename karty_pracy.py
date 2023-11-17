import pandas as pd
import openpyxl
from tkinter import filedialog
import os
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import customtkinter


def karty_pracy_suwnice():
    customtkinter.set_appearance_mode("dark")
    customtkinter.set_default_color_theme("blue")
    root = customtkinter.CTk()

    root.title("Załaduj Obmiar")
    root.geometry("300x170")
    root.resizable(0, 0)

    miesiac = ()
    rok = ()
    def search():
        global miesiac
        global rok
        miesiac = entry_miesiac.get()
        rok = entry_rok.get()
        temp = list(miesiac)
        temp.clear()
        temptwo = list(rok)
        temptwo.clear()
        clear_entry()
        realizer()
        entry_label.destroy()
        new_label = customtkinter.CTkLabel(root, text=("Karty pracy zostały pomyślnie utworzone \n pod nazwą karty_pracy_suwnice"),
                                           text_color=("#2AAA8A", "#2AAA8A"))
        new_label.grid(row=0, column=0, columnspan=1, padx=30, pady=25, sticky="n")

    def clear_entry():
        entry_miesiac.delete(0, 'end')
        entry_rok.delete(0, 'end')

    entry_miesiac = customtkinter.CTkEntry(master=root, width=70, height=35, border_width=1, justify="center")
    entry_miesiac.grid(row=0, column=0, columnspan=1, padx=70, pady=77, sticky="w")

    entry_rok = customtkinter.CTkEntry(master=root, width=70, height=35, border_width=1, justify="center")
    entry_rok.grid(row=0, column=0, columnspan=1, padx=76, pady=77, sticky="e")

    entry_label = customtkinter.CTkLabel(root, text=("Wprowadz miesiąc oraz rok, \na następnie zatwierdź"))
    entry_label.grid(row=0, column=0, columnspan=1, padx=5, pady=25, sticky="n")

    miesiac_label = customtkinter.CTkLabel(root, text=("Miesiac:"))
    miesiac_label.grid(row=0, column=0, columnspan=1, padx=18, pady=47, sticky="w")

    rok_label = customtkinter.CTkLabel(root, text=("Rok:"))
    rok_label.grid(row=0, column=0, columnspan=1, padx=150, pady=47, sticky="e")

    search_button = customtkinter.CTkButton(root, text='Zatwierdz', width=100, height=35, border_width=1,
                                            command=search)
    search_button.grid(row=0, column=0, padx=15, pady=30, sticky="s")

    def realizer():
        global miesiac
        global rok

        path = filedialog.askopenfilename(initialdir="/", title="Załaduj obmiar",
                                          filetypes=(("pliki excel", "*.xlsx"), ("wszystkie pliki", "*")))

        df = pd.read_excel(path, sheet_name=None, header=1)

        df_all = pd.concat(df.values(), ignore_index=True)

        # Grupowanie i filtrowanie danych
        df_data = df_all[["Nr rejestru", "Data", "Opis prac i wykaz materiału", "Cena", "Nr Suwnicy"]]
        df_data = df_data[["Nr rejestru", "Data", "Opis prac i wykaz materiału", "Opis prac i wykaz materiału", "Cena", "Nr Suwnicy"]]
        df_data.columns = ["Nr rejestru", "Data", "Opis prac i wykaz materiału", "Ilosc", "Cena", "Nr Suwnicy"]

        df_data['Data'] = df_data['Data'].replace(regex=[r'r.'], value='')
        df_data['Data'] = df_data['Data'].str.findall(r"^[0-9]{2}-[0-9]{2}-[0-9]{4}").str.join("")
        df_data['Data'] = df_data['Data'].replace(r'^s*$', float('NaN'), regex=True)
        df_data['Data'] = df_data['Data'].fillna(method='ffill')
        df_data['Data'] = pd.to_datetime(df_data['Data'], format='%d-%m-%Y', errors='coerce', utc=False)

        df_data['Nr Suwnicy'] = df_data['Nr Suwnicy'].fillna(method='ffill')

        y = rok
        m = miesiac

        if m == 2:
            df_data = df_data.loc[(df_data['Data'] >= f'{y}-{m}-01') & (df_data['Data'] <= f'{y}-{m}-28')]
        elif m == 0:
            df_data = df_data.loc[df_data['Data'].dt.strftime('%Y') == f'{y}']

        elif m == 1 or m == 3 or m == 5 or m == 7 or m == 8 or m == 10 or m == 12:
            df_data = df_data.loc[(df_data['Data'] >= f'{y}-{m}-01') & (df_data['Data'] <= f'{y}-{m}-31')]
        else:
            df_data = df_data.loc[(df_data['Data'] >= f'{y}-{m}-01') & (df_data['Data'] <= f'{y}-{m}-30')]

        df_data['Data'] = df_data['Data'].dt.date

        # Wyciągnięcie ilości materiałów z tekstu do nowej kolumny
        pattern = r'[-][ ]([0-9]+.{1,}|[0-9])+ (?:szt|kpl|op|kg|mb|l|m|m3|km)'

        df_data['Ilosc'] = df_data['Ilosc'].str.findall(pattern).str.join(", ")
        df_data['Ilosc'] = df_data['Ilosc'].replace(regex=[r','], value='.')
        df_data['Ilosc'] = pd.to_numeric(df_data['Ilosc'], errors='coerce')
        df_data['Cena'] = pd.to_numeric(df_data['Cena'], errors='coerce')

        df_data['Wartosc'] = df_data['Ilosc'] * df_data['Cena']
        df_data['Wartosc'] = df_data['Wartosc'].round(decimals=2)

        df_data = df_data[["Nr rejestru", "Data", "Opis prac i wykaz materiału", "Cena", "Wartosc", "Nr Suwnicy"]]

        df_data['Cena'] = df_data['Cena'].round(decimals=3)

        df_data['Nr rejestru'] = df_data['Nr rejestru'].replace(regex=[r'/[0-9][0-9]'], value='').str.join("")
        df_data['Nr rejestru'] = pd.to_numeric(df_data['Nr rejestru'], errors='coerce')
        df_data['Nr rejestru'] = df_data['Nr rejestru'].round(decimals=0)

        df_data['Opis prac i wykaz materiału'] = df_data['Opis prac i wykaz materiału'].replace(regex=[r'20[0-9][0-9]r.'], value='').str.join("")

        # Usuwanie daty, które służyły głównie do wywołania odpowienich wierszy z excela
        odpowiednia_data = []
        for i, j in zip(df_data['Nr rejestru'], df_data['Data']):
            if i > 0:
                odpowiednia_data.append(j)
            else:
                odpowiednia_data.append(i)

        df_data['Data'] = odpowiednia_data

        # Sortowanie według nr rejerstu i aby całość zachowała sens
        df_data['Sortowanie'] = df_data['Nr rejestru']

        kolejnosc = []
        for i in df_data['Sortowanie']:
            if i > 0:
                kolejnosc.append(i*100)
            else:
                kolejnosc.append(kolejnosc[-1] + 1)

        df_data['Sortowanie'] = kolejnosc
        df_data = df_data.sort_values(by='Sortowanie')

        df_data = df_data.reset_index()
        df_data['index'] = df_data.index

        # dodawanie nr urządzenia przed każdą pracą, k - nr indexu, jedyna możliwość wklejania nowych wierszy do dataframe. -0.5 aby dodało pomiędzy i nie nadpistwało
        nr_urzadzenia = []
        for i, j, k in zip(df_data['Nr rejestru'], df_data['Nr Suwnicy'], df_data['index']):
            if i >= 0:
                print(f"{i} {j} {k}")
                nr_urzadzenia.append(j)
                new_row = ['', '', '', j, '', '', '', '']
                df_data.loc[k-0.5] = new_row
            else:
                pass

        df_data = df_data.sort_index().reset_index(drop=True)

        df_data = df_data[["Nr rejestru", "Data", "Opis prac i wykaz materiału", "Cena", "Wartosc"]]

        nr_rejestrow = []
        for i in df_data['Nr rejestru']:
            nr_rejestrow.append(i)

        if os.path.exists("karta_pracy_suwnice.xlsx"):
            os.remove("karta_pracy_suwnice.xlsx")

        # Utworzenie kraty pracy
        df_data.to_excel("karta_pracy_suwnice.xlsx", index=False)

        # Ustawienie szerokości kolumn, poprawienie czytelności i gotowe do wydruku
        writer = pd.ExcelWriter('karta_pracy_suwnice.xlsx')
        df_data.to_excel(writer, sheet_name='Sheet1', index=False, na_rep='')

        writer.sheets['Sheet1'].set_column(0, 0, 11)
        writer.sheets['Sheet1'].set_column(1, 1, 14)
        writer.sheets['Sheet1'].set_column(2, 2, 85)
        writer.sheets['Sheet1'].set_column(3, 3, 20)
        writer.sheets['Sheet1'].set_column(4, 4, 20)

        writer._save()

        workbook = openpyxl.load_workbook('karta_pracy_suwnice.xlsx')

        worksheet = workbook.active

        # zmiana czcionek i formatowanie tekstu
        font1 = Font(name='Arial', size=14, bold=True, italic=False)
        for row in worksheet['D:E']:
            for cell in row:
                cell.font = font1
                cell.number_format = '#,##0.00 zł'

        for row in worksheet.iter_rows():
            if row[3].value is not None:
               row[2].font = font1

        font3 = Font(name='Arial', size=12, bold=True, italic=False)
        for cell in worksheet['B']:
            cell.font = font3

        font = Font(name='Arial', size=12, bold=True, italic=True)
        for row in worksheet['A1:E1']:
            for cell in row:
                cell.font = font

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        for row in worksheet['A:B']:
            for cell in row:
                cell.fill = PatternFill(start_color="00C0C0C0", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

        font2 = Font(name='Arial CE', size=28, bold=True, italic=False)
        for row in worksheet['A']:
            if row.value in nr_rejestrow:
                row.fill = PatternFill(start_color="00CC99FF", fill_type="solid")
                row.font = font2
            else:
                pass

        # obramowanie
        def set_border(ws, cell_range):
            dashdot = Side(border_style="hair", color="000000")
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=dashdot, left=dashdot, right=dashdot, bottom=dashdot)

        set_border(worksheet, 'A:E')

        for row in worksheet['C']:
            if row.value in nr_urzadzenia:
                row.font = Font(name='Arial CE', size=16, bold=True, italic=False, underline='single')
                row.alignment = Alignment(horizontal='center')
            else:
                pass

        for row in worksheet.iter_rows():
            row[4].border = Border(right=Side(border_style='medium', color='FF000000'), left=Side(border_style="hair", color="000000"), top=Side(border_style="hair", color="000000"), bottom=Side(border_style="hair", color="000000"))
            if row[2].value in nr_urzadzenia:
                row[0].border = Border(top=Side(border_style='medium', color='FF000000'))
                row[1].border = Border(top=Side(border_style='medium', color='FF000000'))
                row[2].border = Border(top=Side(border_style='medium', color='FF000000'))
                row[3].border = Border(top=Side(border_style='medium', color='FF000000'))
                row[4].border = Border(top=Side(border_style='medium', color='FF000000'), right=Side(border_style='medium', color='FF000000'))

        workbook.save('karta_pracy_suwnice.xlsx')
    root.mainloop()
